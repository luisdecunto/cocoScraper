"""
Export layer.
Produces CSV and XLSX price comparison and history reports.
"""

import csv
import logging
import os
from collections import defaultdict

import asyncpg
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

logger = logging.getLogger(__name__)

_GREEN_FILL = PatternFill(fill_type="solid", fgColor="00B050")


async def export_latest(pool: asyncpg.Pool, output_path: str) -> None:
    """Write one row per (product, supplier) with the latest price to a CSV."""
    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    async with pool.acquire() as conn:
        rows = await conn.fetch("""
            SELECT p.name, p.sku, p.supplier, p.category,
                   s.price_unit, s.price_bulk, s.stock, s.scraped_at
            FROM products p
            JOIN price_snapshots s ON s.sku = p.sku AND s.supplier = p.supplier
            WHERE s.scraped_at = (
                SELECT MAX(scraped_at) FROM price_snapshots
                WHERE sku = p.sku AND supplier = p.supplier
            )
            ORDER BY p.supplier, p.name;
        """)

    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["name", "sku", "supplier", "category",
                         "price_unit", "price_bulk", "stock", "scraped_at"])
        for r in rows:
            writer.writerow([r["name"], r["sku"], r["supplier"], r["category"],
                              r["price_unit"], r["price_bulk"], r["stock"], r["scraped_at"]])

    logger.info(f"export_latest: {len(rows)} rows ->{output_path}")


async def export_history(
    pool: asyncpg.Pool,
    sku: str,
    supplier: str,
    output_path: str,
) -> None:
    """Write all snapshots for one (sku, supplier) ordered by date ascending to a CSV."""
    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    async with pool.acquire() as conn:
        rows = await conn.fetch("""
            SELECT s.scraped_at, s.price_unit, s.price_bulk, s.stock,
                   p.name, p.category
            FROM price_snapshots s
            JOIN products p ON p.sku = s.sku AND p.supplier = s.supplier
            WHERE s.sku = $1 AND s.supplier = $2
            ORDER BY s.scraped_at ASC;
        """, sku, supplier)

    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["scraped_at", "price_unit", "price_bulk", "stock", "name", "category"])
        for r in rows:
            writer.writerow([r["scraped_at"], r["price_unit"], r["price_bulk"],
                              r["stock"], r["name"], r["category"]])

    logger.info(f"export_history: {len(rows)} rows ->{output_path}")


async def export_comparison(
    pool: asyncpg.Pool,
    output_csv: str,
    output_xlsx: str,
) -> None:
    """
    Cross-supplier comparison. One row per product name, one column pair per supplier.
    Writes uncertain matches (single-supplier products) to uncertain_matches.csv.
    """
    os.makedirs(os.path.dirname(output_csv) or ".", exist_ok=True)

    async with pool.acquire() as conn:
        rows = await conn.fetch("""
            SELECT p.name, p.sku, p.supplier, p.category,
                   s.price_unit, s.price_bulk, s.scraped_at
            FROM products p
            JOIN price_snapshots s ON s.sku = p.sku AND s.supplier = p.supplier
            WHERE s.scraped_at = (
                SELECT MAX(scraped_at) FROM price_snapshots
                WHERE sku = p.sku AND supplier = p.supplier
            )
            ORDER BY p.name, p.supplier;
        """)

    # Collect all supplier names in sorted order
    suppliers = sorted({r["supplier"] for r in rows})

    # Group by normalized name
    groups: dict[str, list] = defaultdict(list)
    for r in rows:
        groups[r["name"].lower().strip()].append(r)

    certain: list[dict] = []
    uncertain: list[dict] = []

    for norm_name, entries in groups.items():
        found_suppliers = {e["supplier"] for e in entries}
        if len(found_suppliers) < 2:
            uncertain.extend(entries)
            continue

        row_data: dict = {"name": entries[0]["name"], "category": entries[0]["category"]}
        prices: list[float] = []
        for s in suppliers:
            match = next((e for e in entries if e["supplier"] == s), None)
            row_data[f"{s}_unit"] = float(match["price_unit"]) if match and match["price_unit"] is not None else None
            row_data[f"{s}_bulk"] = float(match["price_bulk"]) if match and match["price_bulk"] is not None else None
            if match and match["price_unit"] is not None:
                prices.append((float(match["price_unit"]), s))

        if prices:
            prices.sort(key=lambda x: x[0])
            min_price, min_supplier = prices[0]
            max_price = prices[-1][0]
            row_data["cheapest_supplier"] = min_supplier
            row_data["price_diff_pct"] = (
                round((max_price - min_price) / min_price * 100, 2)
                if min_price > 0 else None
            )
        else:
            row_data["cheapest_supplier"] = None
            row_data["price_diff_pct"] = None

        certain.append(row_data)

    # Build headers
    price_cols = []
    for s in suppliers:
        price_cols += [f"{s}_unit", f"{s}_bulk"]
    headers = ["name", "category"] + price_cols + ["cheapest_supplier", "price_diff_pct"]

    # Write CSV
    with open(output_csv, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=headers, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(certain)

    # Write uncertain matches
    uncertain_path = os.path.join(os.path.dirname(output_csv) or ".", "uncertain_matches.csv")
    with open(uncertain_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["name", "sku", "supplier", "category", "price_unit", "price_bulk", "scraped_at"])
        for e in uncertain:
            writer.writerow([e["name"], e["sku"], e["supplier"], e["category"],
                              e["price_unit"], e["price_bulk"], e["scraped_at"]])

    # Write XLSX
    wb = Workbook()
    ws = wb.active
    ws.title = "Comparison"

    ws.append(headers)
    # Bold header
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for row_data in certain:
        ws.append([row_data.get(h) for h in headers])

    # Green fill on cheapest price cell per row
    unit_col_indices = {s: headers.index(f"{s}_unit") + 1 for s in suppliers}
    cheapest_col_idx = headers.index("cheapest_supplier") + 1

    for row_idx, row_data in enumerate(certain, start=2):
        cheapest = row_data.get("cheapest_supplier")
        if cheapest and cheapest in unit_col_indices:
            ws.cell(row=row_idx, column=unit_col_indices[cheapest]).fill = _GREEN_FILL

    # Auto-width columns
    for col in ws.columns:
        max_len = max((len(str(cell.value)) if cell.value is not None else 0) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_len * 1.2 + 2

    wb.save(output_xlsx)
    logger.info(f"export_comparison: {len(certain)} matched rows, "
                f"{len(uncertain)} uncertain ->{output_csv}, {output_xlsx}")
